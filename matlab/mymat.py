from cProfile import label
from csv import excel
from dataclasses import dataclass
from xml.etree.ElementTree import tostring
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import datetime
import time


excel=load_workbook("DATA_2.xlsx")
sheet = excel["DATA_2"]

excel1=load_workbook("DATc.xlsx")
sheet1 = excel1["DATc"]

time_data=[i.value for i in sheet['C'][1:]]
Current_data =[i.value for i in sheet['A'][1:]]

time_data1=[i.value for i in sheet1['C'][1:]]
Current_data1 =[i.value for i in sheet1['A'][1:]]



# for i in range(10):
#     print(time_data[i])
print("this is the data type\n")

my_day = datetime.date.today()
x_dat =[datetime.datetime.combine(my_day,t) for t in time_data]

my_day1 = datetime.date.today()
x_dat1 =[datetime.datetime.combine(my_day1,t) for t in time_data1]

print(type(x_dat[0]))

#x_dat=[datetime.time(t) for t in time_data]
#


line = plt.plot(x_dat,Current_data)

line2 = plt.plot(x_dat1,Current_data1)


legend = plt.legend(["2A", "4A"],title =" current")

plt.title("ChargCurrentFig")
plt.xlabel("Time(m)")
plt.ylabel("Current(A)")
plt.ylim(0,4)
plt.grid(color = 'r', linestyle = '--', linewidth = 0.5)




plt.show()

