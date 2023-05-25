from cProfile import label
from csv import excel
from dataclasses import dataclass
from xml.etree.ElementTree import tostring
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import datetime
import time
import numpy as np


# excel=load_workbook("Temp.xlsx")
# sheet = excel["GG010002"]
excel=load_workbook("GG010001.xlsx")
sheet = excel["GG010001"]
mylist=[]

for i in range(ord('A'),ord('K')):
    mylist.append([j.value for j in sheet[chr(i)][15:]])

fig, ((ax0,ax1),(ax2,ax3))= plt.subplots(ncols=2,nrows=2, sharey=True, figsize=(9, 3))

# sumdata=[Ambient_dat,PointA_dat,PointA1_dat,PointB_dat,PointB1_dat,PointC_dat,PointC1_dat,PointD_dat,PointD1_dat]
LabelText =["Ambient", "PointA","PointB","PointC","PointD","PointA1","PointB1","PointC1","PointD1"]
dcolor =['red','orange','black','yellow','green','purple','blue','cyan','brown']


ax0.set_title("Points A and A1")
ax0.plot(mylist[0],mylist[1])
ax0.plot(mylist[0],mylist[2])
ax0.annotate(str(max(mylist[2]))+'℃',xy=(mylist[0][mylist[2].index(max(mylist[2]))],max(mylist[2])),xytext=(mylist[0][mylist[2].index(max(mylist[2]))],max(mylist[2])),arrowprops=dict(facecolor='pink', shrink=0.05))
ax0.plot(mylist[0],mylist[6])
ax0.annotate(str(max(mylist[6]))+'℃',xy=(mylist[0][mylist[6].index(max(mylist[6]))],max(mylist[6])),xytext=(mylist[0][mylist[6].index(max(mylist[6]))],max(mylist[6])),arrowprops=dict(facecolor='pink', shrink=0.05))
ax0.grid(color = 'r', linestyle = '--', linewidth = 0.5)
ax0.legend(["Ambient","PointA","PointA1"],title="TestPoints A and A1")  
ax0.set_xlabel("Time(s)")
ax0.set_ylabel("Temperature(℃)")



ax1.set_title("Points B and B1")
ax1.plot(mylist[0],mylist[1])
ax1.plot(mylist[0],mylist[3])
ax1.annotate(str(max(mylist[3]))+'℃',xy=(mylist[0][mylist[3].index(max(mylist[3]))],max(mylist[3])),xytext=(mylist[0][mylist[3].index(max(mylist[3]))],max(mylist[3])),arrowprops=dict(facecolor='pink', shrink=1))

ax1.plot(mylist[0],mylist[7])
ax1.annotate(str(max(mylist[7]))+'℃',xy=(mylist[0][mylist[7].index(max(mylist[7]))],max(mylist[7])),xytext=(mylist[0][mylist[7].index(max(mylist[7]))],max(mylist[7])),arrowprops=dict(facecolor='pink', shrink=0.05))

ax1.grid(color = 'r', linestyle = '--', linewidth = 0.5)
ax1.legend(["Ambient","PointB","PointB1"],title="TestPoints B and B1")
legend = plt.legend(LabelText,title =" Temperature")
ax1.set_xlabel("Time(s)")


ax2.set_title("Points C and C1")
ax2.plot(mylist[0],mylist[1])
ax2.plot(mylist[0],mylist[4])
ax2.annotate(str(max(mylist[4]))+'℃',xy=(mylist[0][mylist[4].index(max(mylist[4]))],max(mylist[4])),xytext=(mylist[0][mylist[4].index(max(mylist[4]))],max(mylist[4])),arrowprops=dict(facecolor='pink', shrink=0.05))

ax2.plot(mylist[0],mylist[8])
ax2.annotate(str(max(mylist[8]))+'℃',xy=(mylist[0][mylist[8].index(max(mylist[8]))],max(mylist[8])),xytext=(mylist[0][mylist[8].index(max(mylist[8]))],max(mylist[8])),arrowprops=dict(facecolor='pink', shrink=0.05))

ax2.grid(color = 'r', linestyle = '--', linewidth = 0.5)
ax2.legend(["Ambient","PointC","PointC1"],title="TestPoints C and C1")
ax2.set_xlabel("Time(s)")
ax2.set_ylabel("Temperature(℃)")

ax3.set_title("Points D and D1")
ax3.plot(mylist[0],mylist[1])
ax3.plot(mylist[0],mylist[5])
ax3.annotate(str(max(mylist[5]))+'℃',xy=(mylist[0][mylist[5].index(max(mylist[5]))],max(mylist[5])),xytext=(mylist[0][mylist[5].index(max(mylist[5]))],max(mylist[5])),arrowprops=dict(facecolor='pink', shrink=0.05))

ax3.plot(mylist[0],mylist[9])
ax3.annotate(str(max(mylist[9]))+'℃',xy=(mylist[0][mylist[9].index(max(mylist[9]))],max(mylist[9])),xytext=(mylist[0][mylist[9].index(max(mylist[9]))],max(mylist[9])),arrowprops=dict(facecolor='pink', shrink=0.05))

ax3.grid(color = 'r', linestyle = '--', linewidth = 0.5)
ax3.legend(["Ambient","PointD","PointD1"],title="TestPoints D and D1")
ax3.set_xlabel("Time(s)")


fig1 = plt.gcf


fig2 = plt.figure()

for i,j in zip(mylist[1:],dcolor):
    plt.annotate(str(max(i))+'℃',xy=(mylist[0][i.index(max(i))],max(i)),xytext=(mylist[0][i.index(max(i))],max(i)))
    plt.plot(mylist[0],i,color=j)
   # plt.plot(time_data[i.index(max(i))],max(i),'+')

legend = plt.legend(LabelText,title =" Temperature")
plt.title("Temperature rise curve")
plt.xlabel("Time(s)")
plt.ylabel("Temperature(℃)")
plt.ylim(0,80)
plt.grid(color = 'r', linestyle = '--', linewidth = 0.5)
plt.show()

