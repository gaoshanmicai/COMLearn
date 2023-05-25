from cProfile import label
from csv import excel
from dataclasses import dataclass
from xml.etree.ElementTree import tostring
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import datetime
import time
import numpy as np
import pandas as pd

excel=load_workbook("test.xlsx")
sheet = excel["DAT00000"]

fig, ax1 = plt.subplots()

# excel1=load_workbook("DATc.xlsx")
# sheet1 = excel1["DATc"]
time_data = np.arange(0,8000*5,5)
# time_data2 =np.arange()
Current_data =[i.value for i in sheet['A'][6:]]
volte_data =  [i.value for i in sheet['B'][6:]]

value = max(Current_data[0:7000])
value1 = max(volte_data[100:7000])

minAval =min(Current_data[100:7000])
minVot = min(volte_data[100:7000])

myindex =Current_data.index(value)
myindex1 =volte_data.index(value1)

minVindex = volte_data.index(minVot)
minAindex = Current_data.index(minAval)



# ptime_data =  [i.value for i in sheet['D'][6:]]


# x_dat =[datetime.datetime.combine(my_day, w,None) for w in ptime_data ] 

# for w in ptime_data:
#     # print(type(w))
#     x_dat.append(datetime.datetime.combine(my_day,w))

# time_data1=[i.value for i in sheet1['C'][1:]]
# Current_data1 =[i.value for i in sheet1['A'][1:]]



color = 'tab:red'
ax1.set_xlabel('Time (s)')
ax1.set_ylabel('Voltage(V)', color=color)
ax1.annotate(str(value1)+'V',xy=(time_data[myindex1],value1),xytext=(time_data[myindex1],value1),arrowprops=dict(facecolor='purple', shrink=1))
ax1.annotate(str(minVot)+'V',xy=(time_data[minVindex],minVot),xytext=(time_data[minVindex],minVot),arrowprops=dict(facecolor='yellow', shrink=1))

ax1.plot(time_data, volte_data, color=color)
ax1.tick_params(axis='y', labelcolor=color)

ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

color = 'tab:blue'
ax2.set_ylabel('Current(A)', color=color)  # we already handled the x-label with ax1
ax2.annotate(str(value)+'A',xy=(time_data[myindex],value),xytext=(time_data[myindex],value),arrowprops=dict(facecolor='purple', shrink=1))
ax2.annotate(str(minAval)+'A',xy=(time_data[minAindex],minAval),xytext=(time_data[minAindex],minAval),arrowprops=dict(facecolor='yellow', shrink=1))

ax2.plot(time_data, Current_data,color=color)
ax2.tick_params(axis='y', labelcolor=color)
ax2.set_title("iPad Charging Test")
fig.tight_layout()  # otherwise the right y-label is slightly clipped

# legend = plt.legend(["Current"],title =" iPad Charging")

# plt.xlabel("Time(m)")
# plt.ylabel("Current(A)")
# plt.ylim(0,2)
plt.grid(color = 'r', linestyle = '--', linewidth = 0.5)

plt.savefig('chargeTest.png')

plt.show()
'''
# for i in range(10):
#     print(time_data[i])

print("this is the data type\n")

my_day = datetime.date.today()
x_dat =[datetime.datetime.combine(my_day,t) for t in time_data]

my_day1 = datetime.date.today()
x_dat1 =[datetime.datetime.combine(my_day1,t) for t in time_data1]

print(type(x_dat[0]))

#x_dat=[datetime.time(t) for t in time_data]
#


line = plt.plot(x_dat,Current_data)

line2 = plt.plot(x_dat1,Current_data1)


legend = plt.legend(["2A", "4A"],title =" current")

plt.title("ChargCurrentFig")
plt.xlabel("Time(m)")
plt.ylabel("Current(A)")
plt.ylim(0,4)
plt.grid(color = 'r', linestyle = '--', linewidth = 0.5)




plt.show()
'''

